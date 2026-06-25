import os, json, tempfile, base64
from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import anthropic
from openpyxl import load_workbook
import pandas as pd

app = FastAPI(title="Fit Map — 포트폴리오 진단 시스템")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="../frontend/static"), name="static")

# ── 멘토 데이터 로드 ──
def load_mentors():
    df = pd.read_csv("../data/mentors.csv").fillna("")
    mentors = []
    for _, row in df.iterrows():
        # 국가 태그 자동 추론
        countries = []
        org = str(row["소속"]).lower()
        if any(k in org for k in ["japan", "tokyo", "도쿄", "일본", "lionice", "라이오니스", "k-startup center"]):
            countries.append("🇯🇵 일본")
        if any(k in org for k in ["singapore", "싱가포르"]):
            countries.append("🇸🇬 싱가포르")
        if any(k in org for k in ["동남아", "인터베스트"]):
            countries.append("🌏 동남아")
        if any(k in str(row["특화산업"]).lower() for k in ["동남아시아"]):
            countries.append("🌏 동남아")
        if not countries:
            countries = ["🌏 글로벌"]

        mentors.append({
            "name": row["이름"],
            "org": row["소속"],
            "title": row["직위"],
            "fields": [f.strip() for f in str(row["특화분야"]).replace("\n", ",").split(",") if f.strip()],
            "industries": [i.strip() for i in str(row["특화산업"]).replace("\n", ",").split(",") if i.strip()],
            "years": int(float(row["경력(년)"])) if row["경력(년)"] else 0,
            "countries": countries,
        })
    return mentors

MENTORS = load_mentors()

def match_mentors(top_fields: list, company_industries: list, top_n: int = 5) -> list:
    scored = []
    for m in MENTORS:
        score = 0
        matched_fields = []
        for i, field in enumerate(top_fields):
            if field in m["fields"]:
                score += (3 - i)
                matched_fields.append(field)
        for ind in company_industries:
            if ind in m["industries"] or "전분야" in m["industries"]:
                score += 2
                break
        if m["years"] >= 20:
            score += 1
        if score > 0:
            scored.append({**m, "score": score, "matched_fields": matched_fields})

    scored.sort(key=lambda x: (-x["score"], -x["years"]))

    # 5명 미달 시 전체 풀에서 보완
    if len(scored) < top_n:
        seen = {m["name"] for m in scored}
        fallback = sorted(
            [m for m in MENTORS if m["name"] not in seen],
            key=lambda x: -x["years"]
        )
        for m in fallback:
            if len(scored) >= top_n:
                break
            scored.append({**m, "score": 0, "matched_fields": []})

    return scored[:top_n]

# ── 밸류업 프로그램 ──
VALUEUP_PROGRAMS = """
1. Private IR (후속투자 강화)
   - 외부 투자자 3~5개사와 포트폴리오사 1:1 소규모 IR
   - 일정: 3~11월 월 1회 / KPI: 총 9회 개최

2. 데모데이 (후속투자 강화)
   - 2026년 7월 / 한국과학기술회관
   - ICT, 로봇, 모빌리티 세션별 IR / KPI: 200명 참석

3. 후속투자자 연계 (투자사 DB 매칭)
   - 섹터/라운드/규모별 투자사 DB 매칭
   - KPI: 신규 투자사 DB 50개 이상, 연결 최소 50회

4. 그로스파트너스 전문가 활용 (필수)
   - 상시 멘토링 (세무/회계/재무, 투자유치, 법률, 해외진출 등)
   - KPI: 상시 66시간, 정기 특강 반기 1회

5. Global Express Program (글로벌 진출)
   - 일본·동남아 현지 파트너 연결 (유통, 마케팅, 법무)
   - KPI: 5개사 지원, 기업별 파트너 연결 4회 이상

6. 고민나눔소
   - 전문가 초청 소규모 세션 (반기 5회차)
   - 예산: 350만원/반기

7. TIPS/서울형TIPS 졸업 지원
   - 협약 종료 예정 기업 밀착 관리

8. 기업간 교류회 / 오픈세미나
   - CTO 교류회, OI 교류회 (반기 1회)
"""

SYSTEM_PROMPT = """당신은 마크앤컴퍼니 투자관리팀 AI 어시스턴트입니다.
포트폴리오사의 재무 데이터를 분석해 Fit Map 진단 결과를 생성합니다.

MPES 평가 기준:
- 유동성(Runway) = 현금보유액 / (월평균비용 - 월평균매출액)
  A: ≥18개월, B: 6~18개월, C: <6개월
- 성장성(매출성장률) = (당기매출 - 전기매출) / 전기매출 × 100
  A: ≥20%, B: 0~20%, C: <0%
- 안정성(자본잠식률): A: 잠식없음, B: 부분잠식(0~50%), C: 완전잠식(≥50%)
- 수익성(영업이익): A: 흑자, B: 손익분기점 근접, C: 적자

반드시 JSON만 반환하고 다른 텍스트는 포함하지 마세요."""

OUTPUT_FORMAT = """{
  "company": "회사명",
  "report_date": "작성기준일",
  "financials": {
    "cash": 현금잔고_숫자_원단위,
    "monthly_burn": 월순유출_숫자_원단위,
    "revenue_by_period": [{"period": "2024", "revenue": 숫자, "cost": 숫자}],
    "quarterly": {
      "revenue_2025": {"Q1": 숫자, "Q2": 숫자, "Q3": 숫자, "Q4": 숫자},
      "cost_2025":    {"Q1": 숫자, "Q2": 숫자, "Q3": 숫자, "Q4": 숫자},
      "revenue_2026_plan": {"Q1": 숫자, "Q2": 숫자, "Q3": 숫자, "Q4": 숫자},
      "cost_2026_plan":    {"Q1": 숫자, "Q2": 숫자, "Q3": 숫자, "Q4": 숫자}
    }
  },
  "mpes": {
    "liquidity":    {"grade": "A/B/C", "value": "계산값", "reason": "근거"},
    "growth":       {"grade": "A/B/C", "value": "계산값", "reason": "근거"},
    "stability":    {"grade": "A/B/C", "value": "계산값", "reason": "근거"},
    "profitability":{"grade": "A/B/C", "value": "계산값", "reason": "근거"},
    "overall": "4글자등급",
    "group": "리스크관리 또는 성장지원 또는 스케일업 또는 안정성장",
    "summary": "종합 코멘트 2~3문장"
  },
  "matched_programs": [
    {"rank": 1, "name": "프로그램명", "priority": "P1/P2/P3",
     "reason": "매칭 근거 (재무 수치 언급)", "kpi": "KPI", "timeline": "즉시/1개월내/3개월내/하반기"}
  ],
  "growth_partner_fields": ["추천분야1", "추천분야2", "추천분야3"],
  "company_industries": ["관련산업1", "관련산업2"],
  "biz_context": {
    "current_status": ["현재 진행 중인 사업 현황 요약 (성과계획 기반)"],
    "risks": ["주요 리스크"],
    "targets": ["타겟 국가/시장"]
  },
  "action_items": ["즉시 액션 1", "액션 2", "액션 3"]
}"""

def parse_excel(filepath: str) -> dict:
    wb = load_workbook(filepath, data_only=True)
    result = {"sheets": {}}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(max_row=60, values_only=True):
            if any(c is not None for c in row):
                rows.append(list(row[:18]))
        result["sheets"][sn] = rows
    return result

def call_claude(client, messages):
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4000,
        system=SYSTEM_PROMPT,
        messages=messages
    )
    raw = msg.content[0].text.strip()
    for tag in ["```json", "```"]:
        if tag in raw:
            raw = raw.split(tag)[1].split("```")[0].strip()
            break
    return json.loads(raw)

def generate_mentor_reasons(client, company_name: str, biz_context: dict, mentors: list, field: str) -> list:
    """각 멘토에 대해 기업 맞춤 매칭 이유 생성"""
    prompt = f"""기업 '{company_name}'에 대해 아래 멘토들이 왜 적합한지 간결하게 설명해주세요.

기업 현황: {json.dumps(biz_context, ensure_ascii=False)}
추천 분야: {field}

멘토 목록:
{json.dumps([{"name": m["name"], "org": m["org"], "title": m["title"], "fields": m["fields"], "years": m["years"]} for m in mentors], ensure_ascii=False)}

아래 JSON만 반환:
[
  {{"name": "멘토명", "expertise": "전문성 한 줄 (25자 이내)", "match_reason": "이 기업 맞춤 매칭 이유 한 줄 (40자 이내)"}},
  ...
]"""

    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = msg.content[0].text.strip()
    for tag in ["```json", "```"]:
        if tag in raw:
            raw = raw.split(tag)[1].split("```")[0].strip()
            break
    reasons = json.loads(raw)
    # 멘토 데이터에 이유 병합
    reason_map = {r["name"]: r for r in reasons}
    for m in mentors:
        r = reason_map.get(m["name"], {})
        m["expertise"] = r.get("expertise", "")
        m["match_reason"] = r.get("match_reason", "")
    return mentors

@app.get("/")
async def root():
    return FileResponse("../frontend/static/index.html")

@app.get("/health")
async def health():
    return {"status": "ok", "mentors": len(MENTORS)}

@app.post("/analyze")
async def analyze(file: UploadFile):
    if not file.filename.endswith((".xlsx", ".xls", ".pdf")):
        raise HTTPException(400, "엑셀(.xlsx) 또는 PDF 파일만 업로드 가능합니다.")

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(500, "ANTHROPIC_API_KEY 환경변수를 설정해주세요.")

    suffix = ".pdf" if file.filename.endswith(".pdf") else ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    try:
        client = anthropic.Anthropic(api_key=api_key)

        if file.filename.endswith(".pdf"):
            with open(tmp_path, "rb") as f:
                pdf_b64 = base64.standard_b64encode(f.read()).decode()
            messages = [{"role": "user", "content": [
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text": f"이 PDF(IR자료/사업계획서)를 분석해 Fit Map 진단 결과를 반환하세요.\n그로스파트너스 분야: B2B영업, 세무/회계/재무, 투자유치(IR), 해외 진출, 홍보마케팅, HR/노무/채용, 법무/법률, 빅데이터/AI\n\n밸류업 프로그램:\n{VALUEUP_PROGRAMS}\n\n출력형식(JSON만):\n{OUTPUT_FORMAT}"}
            ]}]
        else:
            excel_data = parse_excel(tmp_path)
            messages = [{"role": "user", "content": f"""아래 엑셀 데이터를 분석해 Fit Map 진단 결과를 반환하세요.
그로스파트너스 추천 분야는 이 중에서 선택: B2B영업, 세무/회계/재무, 투자유치(IR), 해외 진출, 홍보마케팅, HR/노무/채용, 법무/법률, 빅데이터/AI

엑셀 데이터:
{json.dumps(excel_data, ensure_ascii=False, default=str)[:8000]}

밸류업 프로그램:
{VALUEUP_PROGRAMS}

출력형식(JSON만):
{OUTPUT_FORMAT}"""}]

        result = call_claude(client, messages)

        # ── 멘토 매칭 ──
        top_fields = result.get("growth_partner_fields", [])[:3]
        company_industries = result.get("company_industries", [])
        biz_context = result.get("biz_context", {})

        recommended_mentors = {}
        for field in top_fields:
            mentors_for_field = match_mentors([field], company_industries, top_n=5)
            # AI 매칭 이유 생성
            mentors_with_reasons = generate_mentor_reasons(
                client, result.get("company", ""), biz_context, mentors_for_field, field
            )
            recommended_mentors[field] = mentors_with_reasons

        result["recommended_mentors"] = recommended_mentors
        result["growth_partner_fields"] = top_fields
        return result

    except json.JSONDecodeError as e:
        raise HTTPException(500, f"AI 응답 파싱 오류: {str(e)}")
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        os.unlink(tmp_path)
